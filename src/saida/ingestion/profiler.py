from __future__ import annotations

from pathlib import Path

import duckdb


class DatasetProfiler:
    def profile_tabular(self, path: str) -> dict:
        ext = Path(path).suffix.lower()
        if ext == ".xlsx":
            return self._profile_xlsx(path)
        con = duckdb.connect()
        try:
            escaped = path.replace("'", "''")
            if ext == ".csv":
                relation = con.sql(
                    f"SELECT * FROM read_csv_auto('{escaped}', sample_size=-1, strict_mode=false, header=true)"
                )
            else:
                relation = con.sql(f"SELECT * FROM read_json_auto('{escaped}')")
            columns = [c[0] for c in relation.description]
            count_row = relation.count("*").fetchone()
            count = int(count_row[0]) if count_row is not None else 0
            return {"columns": columns, "row_count": int(count)}
        finally:
            con.close()

    def _profile_xlsx(self, path: str) -> dict:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.worksheets[0] if wb.worksheets else None
            if ws is None:
                return {"columns": [], "row_count": 0}

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return {"columns": [], "row_count": 0}

            header = rows[0]
            columns = [str(v) if v is not None and str(v).strip() else f"column_{i+1}" for i, v in enumerate(header)]
            data_rows = 0
            for row in rows[1:]:
                if any(v is not None and str(v).strip() != "" for v in row):
                    data_rows += 1
            return {"columns": columns, "row_count": data_rows}
        finally:
            wb.close()
