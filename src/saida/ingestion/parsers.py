from __future__ import annotations

import json
from pathlib import Path


TEXT_EXTENSIONS = {".txt", ".md", ".log", ".py", ".sql", ".json", ".csv", ".docx", ".pdf"}
TABULAR_EXTENSIONS = {".csv", ".json", ".xlsx"}


def detect_kind(path: str) -> str:
    ext = Path(path).suffix.lower()
    if ext in TABULAR_EXTENSIONS:
        return "tabular"
    if ext in TEXT_EXTENSIONS:
        return "document"
    if ext in {".pdf", ".docx"}:
        return "document"
    return "binary"


def parse_pdf(path: str) -> str:
    from pypdf import PdfReader

    reader = PdfReader(path)
    chunks: list[str] = []
    for page in reader.pages:
        text = page.extract_text() or ""
        if text.strip():
            chunks.append(text.strip())
    return "\n".join(chunks)


def parse_docx(path: str) -> str:
    from docx import Document

    doc = Document(path)
    chunks: list[str] = []
    for paragraph in doc.paragraphs:
        txt = paragraph.text.strip()
        if txt:
            chunks.append(txt)
    return "\n".join(chunks)


def parse_xlsx(path: str) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    sections: list[str] = []
    for ws in wb.worksheets:
        lines: list[str] = [f"[Sheet: {ws.title}]"]
        for row in ws.iter_rows(values_only=True):
            values = ["" if v is None else str(v) for v in row]
            if any(v != "" for v in values):
                lines.append("\t".join(values))
        if len(lines) > 1:
            sections.append("\n".join(lines))
    wb.close()
    return "\n\n".join(sections)


def parse_text(path: str) -> str:
    p = Path(path)
    ext = p.suffix.lower()
    if ext == ".json":
        raw = json.loads(p.read_text(encoding="utf-8", errors="ignore"))
        return json.dumps(raw, ensure_ascii=True)
    if ext == ".pdf":
        return parse_pdf(path)
    if ext == ".docx":
        return parse_docx(path)
    if ext == ".xlsx":
        return parse_xlsx(path)
    return p.read_text(encoding="utf-8", errors="ignore")


def semantic_summary(text: str, max_len: int = 300) -> str:
    compact = " ".join(text.split())
    return compact[:max_len]
